from flask import Flask, request, jsonify
import openpyxl
import os
import shutil
from openpyxl.styles import Alignment

app = Flask(__name__)

@app.route("/llenar_riesgo", methods=["POST"])
def llenar_riesgo():
    datos = request.get_json()

    actividad = datos.get("actividad", "")
    riesgos = datos.get("riesgos_detectados", "")
    frecuencia = datos.get("frecuencia", "")
    severidad = datos.get("severidad", "")
    impacto = datos.get("impacto", "")
    medidas = datos.get("medidas_control", "")

    # Cargar el archivo desde tmp/
    origen = "tmp/AST_WM.xlsx"
    if not os.path.exists(origen):
        return jsonify({"error": "Archivo AST_WM.xlsx no encontrado en tmp/"}), 500

    wb = openpyxl.load_workbook(origen)
    ws = wb.active

    # Buscar la siguiente fila vacía
    fila = 6
    while ws.cell(row=fila, column=1).value:
        fila += 1

    # Limpieza de texto
    actividad = actividad.replace("*", "")
    riesgos = riesgos.replace("*", "")
    medidas = medidas.replace("*", "")

    # Escribir los datos
    ws.cell(row=fila, column=1, value=actividad)
    ws.cell(row=fila, column=2, value=riesgos)
    ws.cell(row=fila, column=3, value="B" if "B" in riesgos else "M")  # Condiciones seg.
    ws.cell(row=fila, column=4, value=frecuencia)
    ws.cell(row=fila, column=5, value=severidad)
    ws.cell(row=fila, column=6, value=impacto)
    ws.cell(row=fila, column=7, value=medidas)

    # Formato visual
    for col in range(1, 8):
        celda = ws.cell(row=fila, column=col)
        texto = str(celda.value).strip()
        if len(texto) < 30:
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            celda.alignment = Alignment(horizontal="justify", vertical="top", wrap_text=True)

    # Guardar en tmp
    ruta_tmp = "/tmp/AST_WM.xlsx"
    wb.save(ruta_tmp)

    # Copiar a static para descarga
    os.makedirs("static", exist_ok=True)
    shutil.copy(ruta_tmp, "static/AST_WM.xlsx")

    return jsonify({
        "mensaje": f"Análisis registrado en fila {fila}.",
        "fila_insertada": fila
    })

if __name__ == "__main__":
    app.run(debug=True)
